#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""色チェック - 結果をUTF-8ファイルに出力"""
import os, sys
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("pip install openpyxl"); exit(1)

ROOT_FOLDER         = r"C:\Users\USER\Dropbox\②顧客管理表"
DELIVERY_SHEET_NAME = "配達表"
DEL_COL_COUNT  = 5
DEL_COL_NAME   = 3
DEL_COL_TYPE   = 4
OUTPUT          = r"C:\Users\USER\collection-app\color_result.txt"

def resolve_color(fill):
    if fill is None or fill.fill_type in (None, 'none'):
        return None
    fg = fill.fgColor
    if fg is None: return None
    if fg.type == 'rgb':
        rgb = fg.rgb
        if not rgb or rgb in ('00000000',): return None
        color = rgb[2:] if len(rgb) == 8 else rgb
        if color.upper() in ('FFFFFF', '000000'): return None
        return f"#{color.upper()}"
    elif fg.type == 'theme':
        return f"[theme:{fg.theme} tint:{fg.tint:.2f}]"
    elif fg.type == 'indexed':
        return f"[indexed:{fg.indexed}]"
    return None

results = []

for entry in sorted(os.listdir(ROOT_FOLDER)):
    full = os.path.join(ROOT_FOLDER, entry)
    candidates = []
    if os.path.isdir(full):
        for fname in sorted(os.listdir(full)):
            if not fname.startswith('~') and fname.lower().endswith(('.xlsx','.xlsm')):
                candidates.append((os.path.join(full, fname), fname))
    elif os.path.isfile(full) and full.lower().endswith(('.xlsx','.xlsm')):
        candidates.append((full, entry))

    for fpath, fname in candidates:
        try:
            wb = openpyxl.load_workbook(fpath, read_only=False, data_only=True)
        except Exception as e:
            continue
        if DELIVERY_SHEET_NAME not in wb.sheetnames:
            wb.close(); continue
        ws = wb[DELIVERY_SHEET_NAME]
        colors_found = {}
        for row in ws.iter_rows(min_row=2):
            if len(row) <= DEL_COL_COUNT: continue
            c = resolve_color(row[DEL_COL_COUNT].fill)
            if not c: continue
            name = str(row[DEL_COL_NAME].value or '').strip()
            typ  = str(row[DEL_COL_TYPE].value or '').strip()
            if name in ('None','nan',''): continue
            colors_found.setdefault(c, []).append((row[DEL_COL_COUNT].row, name, typ))
        wb.close()
        if colors_found:
            results.append(f"\n[{fname}]")
            for color, entries in sorted(colors_found.items()):
                results.append(f"  色: {color}  ({len(entries)}件)")
                for rn, nm, tp in entries[:8]:
                    results.append(f"    行{rn}: {nm}  種類={tp or '(空)'}")
                if len(entries) > 8:
                    results.append(f"    ... 他{len(entries)-8}件")

with open(OUTPUT, 'w', encoding='utf-8') as f:
    f.write('\n'.join(results))

print(f"完了 → {OUTPUT}")
