#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
配達表シートの「数量」列（F列）のセル背景色を検出して一覧表示するスクリプト
"""

import os, re
from datetime import datetime

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    print("エラー: openpyxl がインストールされていません。pip install openpyxl")
    exit(1)

ROOT_FOLDER         = r"C:\Users\USER\Dropbox\②顧客管理表"
DELIVERY_SHEET_NAME = "配達表"
DEL_COL_COUNT       = 5   # F列（0始まり） = 数量
DEL_COL_NAME        = 3   # D列 = 名前
DEL_COL_TYPE        = 4   # E列 = 種類

THEME_COLOR_NAMES = {
    0: "背景1(白系)", 1: "テキスト1(黒系)", 2: "背景2", 3: "テキスト2",
    4: "アクセント1", 5: "アクセント2", 6: "アクセント3", 7: "アクセント4",
    8: "アクセント5", 9: "アクセント6",
}

def resolve_color(fill):
    """セルのfillからRGB文字列または説明を返す"""
    if fill is None or fill.fill_type in (None, 'none'):
        return None
    fg = fill.fgColor
    if fg is None:
        return None
    if fg.type == 'rgb':
        rgb = fg.rgb  # 例: "FFFF0000"
        if rgb in ('00000000', 'FF000000', None, ''):
            return None
        alpha = rgb[:2]
        color = rgb[2:]
        if color.upper() == 'FFFFFF':
            return None  # 白は無色扱い
        return f"#{color.upper()}"
    elif fg.type == 'theme':
        name = THEME_COLOR_NAMES.get(fg.theme, f"テーマ{fg.theme}")
        tint = f" tint={fg.tint:.2f}" if fg.tint and fg.tint != 0 else ""
        return f"[テーマ色: {name}{tint}]"
    elif fg.type == 'indexed':
        return f"[インデックス色: {fg.indexed}]"
    return None

def main():
    print("=" * 60)
    print("配達表 数量列（F列）セル色チェック")
    print("=" * 60)

    found_any = False

    for entry in sorted(os.listdir(ROOT_FOLDER)):
        full = os.path.join(ROOT_FOLDER, entry)

        candidates = []
        if os.path.isdir(full):
            for fname in sorted(os.listdir(full)):
                if not fname.startswith('~') and fname.lower().endswith(('.xlsx', '.xlsm')):
                    candidates.append(os.path.join(full, fname))
        elif os.path.isfile(full) and full.lower().endswith(('.xlsx', '.xlsm')):
            candidates.append(full)

        for fpath in candidates:
            fname = os.path.basename(fpath)
            try:
                # read_only=False で開かないと色情報が取れない
                wb = openpyxl.load_workbook(fpath, read_only=False, data_only=True)
            except Exception as e:
                print(f"  [エラー] {fname}: {e}")
                continue

            if DELIVERY_SHEET_NAME not in wb.sheetnames:
                wb.close()
                continue

            ws = wb[DELIVERY_SHEET_NAME]
            colors_found = {}  # color_str -> [(row, name, type)]

            for row in ws.iter_rows(min_row=2):
                if len(row) <= DEL_COL_COUNT:
                    continue
                count_cell = row[DEL_COL_COUNT]
                color = resolve_color(count_cell.fill)
                if color:
                    name_val = row[DEL_COL_NAME].value if len(row) > DEL_COL_NAME else ''
                    type_val = row[DEL_COL_TYPE].value if len(row) > DEL_COL_TYPE else ''
                    name_str = str(name_val).strip() if name_val else ''
                    type_str = str(type_val).strip() if type_val else ''
                    if name_str not in ('None', 'nan', ''):
                        if color not in colors_found:
                            colors_found[color] = []
                        colors_found[color].append((count_cell.row, name_str, type_str))

            wb.close()

            if colors_found:
                found_any = True
                print(f"\n【{fname}】")
                for color, entries in sorted(colors_found.items()):
                    print(f"  色: {color}  ({len(entries)}件)")
                    for row_num, name, typ in entries[:5]:
                        print(f"    行{row_num}: {name}  種類={typ or '(空)'}")
                    if len(entries) > 5:
                        print(f"    ... 他{len(entries)-5}件")

    if not found_any:
        print("\n色付きセルは見つかりませんでした。")
        print("（テーマ色・透明・白は除外しています）")

    print("\n完了")

if __name__ == '__main__':
    main()
