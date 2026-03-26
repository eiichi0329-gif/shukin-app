#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
配達用アプリ データ更新スクリプト（6店舗対応版）
====================================================
Dropbox 内の各店舗フォルダにある Excel ファイルを読み込み、
collection-app/data.js を生成します。

必要ライブラリ:
    pip install openpyxl

使い方:
    python update_app.py
"""

import os, json, re
from datetime import datetime
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    print("エラー: openpyxl がインストールされていません。")
    print("  pip install openpyxl  を実行してください。")
    exit(1)

# ──────────────────────────────────────────────────────────────
# 設定
# ──────────────────────────────────────────────────────────────
ROOT_FOLDER    = r"C:\Users\USER\Dropbox\②顧客管理表"
OUTPUT_FILE    = r"C:\Users\USER\collection-app\data.js"
SHEET_NAME          = "顧客リスト"
DELIVERY_SHEET_NAME = "配達表"
MIN_DATA_MONTH      = "2026-02"          # 2026年2月分以降を対象

# 配達表を有効にするか（False=タブ非表示, True=タブ表示）
FEATURE_DELIVERY_ENABLED = True

# GAS ウェブアプリ URL（設定するとどの端末からでも自動で同期されます）
# 空文字のままにするとアプリ内の「連携設定」画面で端末ごとに設定できます
GAS_URL        = "https://script.google.com/macros/s/AKfycbzaN9MnPONOoSH8DzrjxZhYBPstdmWsxnG217x8HdE0CZ9Oxckzxx55RVCqDZgMcTQ/exec"

# 店舗キーワード → 正式店舗名
# フォルダ名・ファイル名のどちらに含まれていても識別します
STORE_KEYWORDS = {
    '下関':  '下関店',
    '北九州': '北九州店',
    '宇部':  '宇部店',
    '宗像':  '宗像店',
    '飯塚':  '飯塚店',
    '福岡東': '福岡東店',
}

# 顧客リスト Excel 列インデックス（0始まり）
COL_CODE    = 2   # C列: 顧客コード
COL_NAME    = 3   # D列: 名前
COL_PAYMENT = 8   # I列: 口座振替番号（数字あり=口座振替 / 空=現金）
COL_ADDRESS = 10  # K列: 住所
COL_AMOUNT  = 66  # BO列: 集金金額

# 配達表 Excel 列インデックス（0始まり）
DEL_COL_CODE      = 2   # C列: コード（ルート計算用）
DEL_COL_NAME      = 3   # D列: 名前
DEL_COL_TYPE      = 4   # E列: 種類
DEL_COL_COUNT     = 5   # F列: 数
DEL_COL_WEEKLY    = 6   # G列: 週間配達予定
# H列 (index 7) はスキップ
DEL_COL_PAYMENT   = 8   # I列: 支払（数字あり=口座振替 / 空=現金）
DEL_COL_VESSEL    = 9   # J列: 容器
DEL_COL_ADDRESS   = 10  # K列: 住所
DEL_COL_NOTES     = 11  # L列: 注意事項
DEL_COL_PHONE     = 12  # M列: 電話番号
DEL_COL_EMERGENCY = 13  # N列: 緊急連絡先
DEL_COL_MEMO      = 14  # O列: 備考
DEL_COL_ABSENT    = 15  # P列: 不在時の対応

# 数量セルの背景色 → 表示ラベルのマッピング（空文字=表示しない）
COUNT_COLOR_MAP = {
    '#FFFF00':         '集金',         # 黄色: 集金
    '#FF0000':         '翌週注文確認',  # 赤
    'theme:8:0.60':    '新規',         # 水色
    'theme:9:0.60':    '再注文',       # ベージュ
}

# データ行としてスキップする名前
SKIP_NAMES = {'None', 'nan', '名前', '顧客名', '氏名', 'お客様名', 'Name', '担当者名', ''}

# ──────────────────────────────────────────────────────────────
# ルート番号計算（コード1〜999→ルート1, 1000〜1999→ルート2, 2000〜2999→ルート3 …）
# ──────────────────────────────────────────────────────────────
def code_to_route(code):
    try:
        c = int(code)
        if c <= 0:
            return 0
        if c < 1000:
            return 1
        return (c - 1000) // 1000 + 2
    except (ValueError, TypeError):
        return 0

# ──────────────────────────────────────────────────────────────
# ファイル名 or 更新日時からデータ月を推定
# ──────────────────────────────────────────────────────────────
def guess_data_month(filename, mtime):
    patterns = [
        r'(\d{4})年\s*(\d{1,2})\s*月',
        r'[\(\（](\d{4})[\s_\-\.\u00b7](\d{1,2})[\)\）]',  # (2026 03) / (2026.3) 等
        r'(\d{4})[_\-\.](\d{1,2})(?!\d)',
        r'(\d{4})(\d{2})(?!\d)',
    ]
    for pat in patterns:
        m = re.search(pat, filename)
        if m:
            y, mo = int(m.group(1)), int(m.group(2))
            if 2020 <= y <= 2035 and 1 <= mo <= 12:
                return f"{y}-{mo:02d}"
    return mtime.strftime('%Y-%m')

# ──────────────────────────────────────────────────────────────
# 文字列から店舗名を識別
# ──────────────────────────────────────────────────────────────
def detect_store(text):
    """フォルダ名やファイル名から店舗名を返す。見つからなければ None。"""
    for keyword, store_name in STORE_KEYWORDS.items():
        if keyword in text:
            return store_name
    return None

# ──────────────────────────────────────────────────────────────
# スキャン対象ファイルの収集
# ROOT_FOLDER 直下のサブフォルダ（店舗別）と
# ROOT_FOLDER 直下の Excel ファイルを両方スキャンします。
# ──────────────────────────────────────────────────────────────
def collect_excel_files(root):
    """
    Returns list of dicts:
      { path, name, mtime, dataMonth, store }
    MIN_DATA_MONTH より前の月のファイルは除外します。
    """
    if not os.path.exists(root):
        print(f"[エラー] フォルダが見つかりません: {root}")
        return []

    result = []

    def add_file(fpath, fname, store_hint=None):
        try:
            mtime = datetime.fromtimestamp(os.path.getmtime(fpath))
        except OSError:
            return
        data_month = guess_data_month(fname, mtime)
        if data_month < MIN_DATA_MONTH:
            return
        # 店舗はファイル名 → フォルダ名 → store_hint の順で判定
        store = detect_store(fname) or (detect_store(store_hint) if store_hint else None)
        if not store:
            print(f"  [スキップ] 店舗名を識別できません: {fname}")
            return
        result.append({
            'path':      fpath,
            'name':      fname,
            'mtime':     mtime,
            'dataMonth': data_month,
            'store':     store,
        })

    for entry in sorted(os.listdir(root)):
        full = os.path.join(root, entry)

        if os.path.isdir(full):
            # サブフォルダをスキャン（店舗別フォルダを想定）
            store_from_folder = detect_store(entry)
            for fname in sorted(os.listdir(full)):
                if fname.startswith(('~', '.')):
                    continue
                if not fname.lower().endswith(('.xlsx', '.xls', '.xlsm')):
                    continue
                add_file(os.path.join(full, fname), fname, store_hint=entry)

        elif os.path.isfile(full):
            # ROOT 直下の Excel ファイル
            if entry.startswith(('~', '.')):
                continue
            if not entry.lower().endswith(('.xlsx', '.xls', '.xlsm')):
                continue
            add_file(full, entry)

    return sorted(result, key=lambda x: (x['store'], x['dataMonth'], x['name']))

# ──────────────────────────────────────────────────────────────
# Excel からレコード抽出
# ──────────────────────────────────────────────────────────────
def extract_from_excel(file_info):
    fpath      = file_info['path']
    fname      = file_info['name']
    data_month = file_info['dataMonth']
    store      = file_info['store']

    try:
        wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
    except Exception as e:
        print(f"    [エラー] 開けませんでした: {e}")
        return []

    if SHEET_NAME not in wb.sheetnames:
        print(f"    [警告] シート '{SHEET_NAME}' が見つかりません")
        print(f"    利用可能シート: {wb.sheetnames}")
        wb.close()
        return []

    ws = wb[SHEET_NAME]
    records = []
    min_cols = max(COL_CODE, COL_NAME, COL_PAYMENT, COL_ADDRESS, COL_AMOUNT) + 1

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < min_cols:
            continue

        name_raw = row[COL_NAME]
        name = str(name_raw).strip() if name_raw is not None else ''
        if not name or name in SKIP_NAMES:
            continue

        # コード / ルート
        try:
            code = int(row[COL_CODE]) if row[COL_CODE] is not None else 0
        except (ValueError, TypeError):
            code = 0
        route = code_to_route(code)

        # 支払方法
        pay_raw = row[COL_PAYMENT]
        is_bank = pay_raw is not None and str(pay_raw).strip() not in ('', 'None', '0', 'nan')

        # 住所
        addr_raw = row[COL_ADDRESS]
        address = str(addr_raw).strip() if addr_raw is not None else ''
        if address in ('None', 'nan'):
            address = ''

        # 金額
        try:
            amount = int(float(str(row[COL_AMOUNT]))) if row[COL_AMOUNT] is not None else 0
        except (ValueError, TypeError):
            amount = 0

        records.append({
            'store':       store,
            'code':        code,
            'route':       route,
            'name':        name,
            'paymentType': 'bank' if is_bank else 'cash',
            'address':     address,
            'amount':      amount,
            'dataMonth':   data_month,
            'sourceFile':  fname,
        })

    wb.close()
    return records

# ──────────────────────────────────────────────────────────────
# セル背景色 → ラベル変換
# ──────────────────────────────────────────────────────────────
def get_count_color_label(cell):
    try:
        fill = cell.fill
        if fill is None or fill.fill_type in (None, 'none'):
            return ''
        fg = fill.fgColor
        if fg is None:
            return ''
        if fg.type == 'rgb':
            rgb = fg.rgb or ''
            if rgb == '00000000':
                return ''
            color = f"#{rgb[2:].upper()}" if len(rgb) == 8 else f"#{rgb.upper()}"
            if color == '#FFFFFF':
                return ''
        elif fg.type == 'theme':
            color = f"theme:{fg.theme}:{fg.tint:.2f}"
        else:
            return ''
        return COUNT_COLOR_MAP.get(color, '')
    except Exception:
        return ''

# ──────────────────────────────────────────────────────────────
# 配達表シートからレコード抽出（今月分ファイルのみ対象）
# ──────────────────────────────────────────────────────────────
def extract_delivery_from_excel(file_info):
    fpath      = file_info['path']
    fname      = file_info['name']
    data_month = file_info['dataMonth']
    store      = file_info['store']

    try:
        # read_only=False にしないとセルの色情報が読み取れない
        wb = openpyxl.load_workbook(fpath, read_only=False, data_only=True)
    except Exception as e:
        print(f"    [エラー] 開けませんでした: {e}")
        return []

    if DELIVERY_SHEET_NAME not in wb.sheetnames:
        print(f"    [警告] シート '{DELIVERY_SHEET_NAME}' が見つかりません（配達表）")
        wb.close()
        return []

    ws = wb[DELIVERY_SHEET_NAME]
    records = []

    def safe_str(val):
        if val is None:
            return ''
        s = str(val).strip()
        return '' if s in ('None', 'nan') else s

    def gcol(row, idx):
        return safe_str(row[idx].value) if len(row) > idx else ''

    for row in ws.iter_rows(min_row=2):
        if not row or len(row) <= DEL_COL_NAME:
            continue
        name = safe_str(row[DEL_COL_NAME].value)
        if not name or name in SKIP_NAMES:
            continue

        code_raw = row[DEL_COL_CODE].value if len(row) > DEL_COL_CODE else None
        try:
            code = int(code_raw) if code_raw is not None else 0
        except (ValueError, TypeError):
            code = 0
        route = code_to_route(code)

        pay_raw = row[DEL_COL_PAYMENT].value if len(row) > DEL_COL_PAYMENT else None
        is_bank = pay_raw is not None and safe_str(pay_raw) not in ('', '0')

        count_label = get_count_color_label(row[DEL_COL_COUNT]) if len(row) > DEL_COL_COUNT else ''

        records.append({
            'store':       store,
            'dataMonth':   data_month,
            'code':        code,
            'route':       route,
            'name':        name,
            'type':        gcol(row, DEL_COL_TYPE),
            'count':       gcol(row, DEL_COL_COUNT),
            'countLabel':  count_label,
            'weekly':      gcol(row, DEL_COL_WEEKLY),
            'paymentType': 'bank' if is_bank else 'cash',
            'vessel':      gcol(row, DEL_COL_VESSEL),
            'address':     gcol(row, DEL_COL_ADDRESS),
            'notes':       gcol(row, DEL_COL_NOTES),
            'phone':       gcol(row, DEL_COL_PHONE),
            'emergency':   gcol(row, DEL_COL_EMERGENCY),
            'memo':        gcol(row, DEL_COL_MEMO),
            'absent':      gcol(row, DEL_COL_ABSENT),
        })

    wb.close()
    return records

# ──────────────────────────────────────────────────────────────
# 名寄せ（同一店舗・同一月・同一コード・同一名前の金額を合算）
# ──────────────────────────────────────────────────────────────
def merge_records(all_records):
    merged = {}
    for rec in all_records:
        key = (rec['store'], rec['dataMonth'], rec['code'], rec['name'])
        sf  = rec.pop('sourceFile', '')
        if key not in merged:
            merged[key] = {**rec, 'sourceFiles': [sf] if sf else []}
        else:
            merged[key]['amount'] += rec['amount']
            if sf and sf not in merged[key]['sourceFiles']:
                merged[key]['sourceFiles'].append(sf)

    result = sorted(
        merged.values(),
        key=lambda x: (x['store'], x['dataMonth'], x['code'], x['name'])
    )
    for i, rec in enumerate(result, 1):
        rec['seq'] = i
    return result

# ──────────────────────────────────────────────────────────────
# メイン
# ──────────────────────────────────────────────────────────────
def main():
    print("=" * 60)
    print("配達用アプリ データ更新スクリプト（6店舗対応版）")
    print("=" * 60)
    print(f"スキャン先  : {ROOT_FOLDER}")
    print(f"対象期間    : {MIN_DATA_MONTH} 以降")
    print(f"対象店舗    : {', '.join(STORE_KEYWORDS.values())}")
    print()

    # ── ファイル収集 ──
    files = collect_excel_files(ROOT_FOLDER)
    if not files:
        print("対象ファイルが見つかりませんでした。")
        print("フォルダパスと店舗フォルダ名を確認してください。")
        return

    print(f"対象ファイル: {len(files)} 件")
    cur_store = None
    for f in files:
        if f['store'] != cur_store:
            cur_store = f['store']
            print(f"  ── {cur_store} ──")
        print(f"    [{f['dataMonth']}] {f['name']}  ({f['mtime'].strftime('%Y-%m-%d')})")
    print()

    # ── 抽出 ──
    all_records = []
    source_meta = []

    for fi in files:
        print(f"処理中: [{fi['store']}] {fi['name']}")
        recs = extract_from_excel(fi)
        all_records.extend(recs)
        print(f"  → {len(recs)} 件抽出")
        source_meta.append({
            'store':     fi['store'],
            'name':      fi['name'],
            'dataMonth': fi['dataMonth'],
            'mtime':     fi['mtime'].strftime('%Y-%m-%d'),
            'count':     len(recs),
        })

    # ── 配達表データ抽出（今月分のみ）──
    current_month = datetime.now().strftime('%Y-%m')
    current_month_files = [f for f in files if f['dataMonth'] == current_month]
    print(f"配達表対象ファイル（{current_month}分）: {len(current_month_files)} 件")

    all_delivery = []
    for fi in current_month_files:
        print(f"配達表処理中: [{fi['store']}] {fi['name']}")
        drecs = extract_delivery_from_excel(fi)
        all_delivery.extend(drecs)
        print(f"  → 配達表 {len(drecs)} 件抽出")

    all_delivery.sort(key=lambda x: (x['store'], x['route'], x['code'], x['name']))
    print(f"配達表合計: {len(all_delivery)} 件")
    print()

    print()
    print(f"抽出合計: {len(all_records)} 件")

    # ── 名寄せ ──
    merged = merge_records(all_records)
    print(f"名寄せ後: {len(merged)} 件")
    print()

    # ── 集計表示 ──
    by_store_month = defaultdict(lambda: {'cnt': 0, 'total': 0})
    for r in merged:
        k = (r['store'], r['dataMonth'])
        by_store_month[k]['cnt']   += 1
        by_store_month[k]['total'] += r['amount']

    print("【店舗×月別集計】")
    prev_store = None
    store_totals = defaultdict(int)
    for (store, month), d in sorted(by_store_month.items()):
        if store != prev_store:
            if prev_store:
                print(f"    小計: {store_totals[prev_store]:,}円")
            print(f"  {store}")
            prev_store = store
        print(f"    {month}分: {d['cnt']:>4}件  {d['total']:>10,}円")
        store_totals[store] += d['total']
    if prev_store:
        print(f"    小計: {store_totals[prev_store]:,}円")

    grand_total = sum(r['amount'] for r in merged)
    print(f"  ──────────────────────────────")
    print(f"  合計 {len(merged):>5}件  {grand_total:>12,}円")
    print()

    # ── data.js 生成 ──
    meta = {
        'generatedAt':  datetime.now().isoformat(),
        'stores':       list(dict.fromkeys(STORE_KEYWORDS.values())),  # 定義順を保持
        'sourceFiles':  source_meta,
        'totalRecords': len(merged),
        'totalAmount':  grand_total,
    }
    gas_url_line     = f'window.GAS_URL = {json.dumps(GAS_URL)};\n\n' if GAS_URL else ''
    delivery_enabled = 'true' if FEATURE_DELIVERY_ENABLED else 'false'
    js_out = (
        "// ══════════════════════════════════════════════════════\n"
        f"// 自動生成: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
        f"// 件数: {len(merged)} 件  合計: {grand_total:,}円\n"
        "// ══════════════════════════════════════════════════════\n"
        f"{gas_url_line}"
        f"window.FEATURE_DELIVERY_ENABLED = {delivery_enabled};\n\n"
        f"window.COLLECTION_DATA = {json.dumps(merged, ensure_ascii=False, indent=2)};\n\n"
        f"window.DELIVERY_DATA = {json.dumps(all_delivery, ensure_ascii=False, indent=2)};\n\n"
        f"window.DATA_META = {json.dumps(meta, ensure_ascii=False, indent=2)};\n"
    )

    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        f.write(js_out)

    print(f"出力完了: {OUTPUT_FILE}")
    print()
    print("次の手順:")
    print("  1. collection-app/index.html をブラウザで開いてください")
    print("  2. データが自動的に読み込まれます")

if __name__ == '__main__':
    main()
